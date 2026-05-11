"""Generate an Excel workbook from an AuditResult.

Five sheets optimized for consultants who work in spreadsheets daily:
1. **Synthèse**   — cover block: domain, global score, verdict, counters, date.
2. **Scores**     — 6-axis table with colour-coded scores (conditional formatting).
3. **Findings**   — flat, filterable list of every finding across axes.
4. **Pages**      — page-by-page table (title/H1/meta lengths, keywords, reco).
5. **À créer**    — missing strategic pages with traffic estimate and priority.

Design goals:
- Headers in bold with a coloured band so the file looks professional when
  forwarded to a client.
- Wrap long text, freeze header rows, auto-filter on large tables.
- Tabular numbers (titleLength, metaLength, scores) actually formatted as int.
- No hard dependency on fonts or images — pure openpyxl primitives.
"""

from __future__ import annotations

import io
import logging
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from api.models import (
    AgencyBranding,
    AuditResult,
    Finding,
    MissingPage,
    PageAnalysis,
    SectionResult,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Palette — matches the web/PDF so the Excel feels part of the same report.

_HEADER_FILL = PatternFill("solid", fgColor="18181B")
_HEADER_FONT = Font(name="Inter", size=11, bold=True, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="F4F4F5")
_SECTION_FONT = Font(name="Inter", size=10, bold=True, color="52525B")
_BODY_FONT = Font(name="Inter", size=10, color="18181B")
_MUTED_FONT = Font(name="Inter", size=10, color="52525B")
_MONO_FONT = Font(name="JetBrains Mono", size=9, color="52525B")

_BORDER_SIDE = Side(style="thin", color="E4E4E7")
_THIN_BORDER = Border(left=_BORDER_SIDE, right=_BORDER_SIDE, top=_BORDER_SIDE, bottom=_BORDER_SIDE)

_SCORE_FILLS = {
    "critical": PatternFill("solid", fgColor="FEE2E2"),
    "warning": PatternFill("solid", fgColor="FEF3C7"),
    "info": PatternFill("solid", fgColor="DBEAFE"),
    "ok": PatternFill("solid", fgColor="D1FAE5"),
}

_SEVERITY_FILL = {
    "critical": PatternFill("solid", fgColor="FEE2E2"),
    "warning": PatternFill("solid", fgColor="FEF3C7"),
    "info": PatternFill("solid", fgColor="DBEAFE"),
    "ok": PatternFill("solid", fgColor="D1FAE5"),
    "missing": PatternFill("solid", fgColor="EDE9FE"),
}

_PRIORITY_FILL = {
    "high": PatternFill("solid", fgColor="FEE2E2"),
    "medium": PatternFill("solid", fgColor="FEF3C7"),
    "low": PatternFill("solid", fgColor="DBEAFE"),
}

SECTION_LABEL = {
    "security": "Sécurité",
    "seo": "SEO",
    "ux": "UX",
    "content": "Contenu",
    "performance": "Performance",
    "business": "Business",
}


# ---------------------------------------------------------------------------
# Public API


def generate_xlsx(
    audit: AuditResult, *, branding: Optional[AgencyBranding] = None,
) -> bytes:
    wb = Workbook()

    # Workbook-level metadata so the file looks legit when opened by the client
    wb.properties.title = f"Audit {audit.domain}"
    wb.properties.creator = (branding.name if branding and branding.name else "Audit Web IA")
    wb.properties.subject = "Rapport d'audit web"

    # Default first sheet → rename for summary
    summary_ws = wb.active
    summary_ws.title = "Synthèse"
    _write_summary(summary_ws, audit, branding)

    scores_ws = wb.create_sheet("Scores")
    _write_scores(scores_ws, audit)

    findings_ws = wb.create_sheet("Findings")
    _write_findings(findings_ws, audit)

    if audit.pages:
        pages_ws = wb.create_sheet("Pages")
        _write_pages(pages_ws, audit.pages)

    if audit.technicalCrawl and audit.technicalCrawl.rows:
        crawl_ws = wb.create_sheet("Crawl technique")
        _write_technical_crawl(crawl_ws, audit.technicalCrawl)

    if audit.accessibilityAudit and audit.accessibilityAudit.pageScores:
        a11y_ws = wb.create_sheet("Accessibilité")
        _write_a11y(a11y_ws, audit.accessibilityAudit)

    if audit.missingPages:
        missing_ws = wb.create_sheet("À créer")
        _write_missing(missing_ws, audit.missingPages)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Sheets


def _write_summary(
    ws: Worksheet, audit: AuditResult, branding: Optional[AgencyBranding],
) -> None:
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 70
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Rapport d'audit web"
    ws["A1"].font = Font(name="Inter", size=18, bold=True, color="18181B")
    ws.merge_cells("A1:B1")

    agency = branding.name if branding and branding.name else "Audit Web IA"
    ws["A2"] = f"Produit par {agency}"
    ws["A2"].font = _MUTED_FONT
    ws.merge_cells("A2:B2")

    rows = [
        ("Domaine", audit.domain),
        ("URL", audit.url),
        ("Date", audit.createdAt[:10]),
        ("Score global", f"{audit.globalScore}/100"),
        ("Verdict", audit.globalVerdict),
        ("Points critiques", audit.criticalCount),
        ("Avertissements", audit.warningCount),
        ("Pages analysées", len(audit.pages) if audit.pages else 0),
        ("Pages à créer", len(audit.missingPages) if audit.missingPages else 0),
        ("Quick wins", len(audit.quickWins)),
    ]
    row = 4
    for label, value in rows:
        ws.cell(row=row, column=1, value=label).font = _SECTION_FONT
        ws.cell(row=row, column=2, value=value).font = _BODY_FONT
        row += 1

    # Quick wins list
    row += 1
    ws.cell(row=row, column=1, value="Quick wins prioritaires").font = Font(
        name="Inter", size=12, bold=True, color="18181B"
    )
    row += 1
    for i, win in enumerate(audit.quickWins, 1):
        ws.cell(row=row, column=1, value=f"{i:02d}").font = _MONO_FONT
        c = ws.cell(row=row, column=2, value=win)
        c.font = _BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(18, 14 * (len(win) // 80 + 1))
        row += 1


def _write_scores(ws: Worksheet, audit: AuditResult) -> None:
    headers = ["Axe", "Score", "Progression", "Verdict", "# Findings"]
    _write_header(ws, headers)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 14

    by_section = {s.section: s for s in audit.sections}
    for axis in ("security", "seo", "ux", "content", "performance", "business"):
        s = by_section.get(axis)
        label = SECTION_LABEL.get(axis, axis)
        score = s.score if s else audit.scores.get(axis, 0)  # type: ignore[arg-type]
        verdict = s.verdict if s else ""
        findings_count = len(s.findings) if s else 0

        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=label).font = _BODY_FONT
        score_cell = ws.cell(row=row, column=2, value=score)
        score_cell.font = Font(name="Inter", size=11, bold=True, color="18181B")
        score_cell.alignment = Alignment(horizontal="center")
        score_cell.fill = _score_fill(score)
        score_cell.number_format = "0"

        bar = ws.cell(row=row, column=3, value=_progress_bar(score))
        bar.font = _MONO_FONT
        verdict_cell = ws.cell(row=row, column=4, value=verdict)
        verdict_cell.font = _MUTED_FONT
        verdict_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=5, value=findings_count).font = _MONO_FONT
        ws.row_dimensions[row].height = max(20, 14 * (len(verdict) // 80 + 1))

    ws.freeze_panes = "A2"


def _write_findings(ws: Worksheet, audit: AuditResult) -> None:
    headers = [
        "Axe", "Sévérité", "Impact", "Effort", "Titre", "Description",
        "Recommandation", "Actions", "Référence",
    ]
    _write_header(ws, headers)
    for letter, width in zip(
        ("A", "B", "C", "D", "E", "F", "G", "H", "I"),
        (14, 12, 10, 10, 50, 70, 70, 70, 30),
    ):
        ws.column_dimensions[letter].width = width

    row = 2
    for section in audit.sections:
        for f in section.findings:
            ws.cell(row=row, column=1, value=SECTION_LABEL.get(section.section, section.section)).font = _BODY_FONT
            sev_cell = ws.cell(row=row, column=2, value=f.severity.upper())
            sev_cell.font = Font(name="Inter", size=10, bold=True, color="18181B")
            sev_cell.fill = _SEVERITY_FILL.get(f.severity, PatternFill())
            sev_cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=(f.impact or "")).font = _MUTED_FONT
            ws.cell(row=row, column=4, value=(f.effort or "")).font = _MUTED_FONT
            _wrap_cell(ws.cell(row=row, column=5, value=f.title))
            _wrap_cell(ws.cell(row=row, column=6, value=f.description))
            _wrap_cell(ws.cell(row=row, column=7, value=f.recommendation or ""))
            _wrap_cell(ws.cell(row=row, column=8, value="\n".join(f.actions or [])))
            ref = ws.cell(row=row, column=9, value=f.reference or "")
            ref.font = _MONO_FONT
            row += 1

    if row > 2:
        ws.auto_filter.ref = f"A1:I{row - 1}"
    ws.freeze_panes = "A2"


def _write_pages(ws: Worksheet, pages: list[PageAnalysis]) -> None:
    headers = [
        "URL", "Statut", "Title", "Title (car.)", "H1", "Meta description",
        "Meta (car.)", "KW cibles", "KW présents", "KW absents", "Findings",
        "URL recommandée", "Title recommandé",
    ]
    _write_header(ws, headers)
    widths = (55, 12, 50, 10, 40, 60, 10, 35, 35, 35, 60, 45, 50)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2
    for p in pages:
        ws.cell(row=row, column=1, value=p.url).font = _MONO_FONT
        status_cell = ws.cell(row=row, column=2, value=p.status.upper())
        status_cell.fill = _score_fill_for_status(p.status)
        status_cell.font = Font(name="Inter", size=10, bold=True, color="18181B")
        status_cell.alignment = Alignment(horizontal="center")
        _wrap_cell(ws.cell(row=row, column=3, value=p.title or ""))
        ws.cell(row=row, column=4, value=p.titleLength).number_format = "0"
        _wrap_cell(ws.cell(row=row, column=5, value=p.h1 or ""))
        _wrap_cell(ws.cell(row=row, column=6, value=p.metaDescription or ""))
        ws.cell(row=row, column=7, value=p.metaLength).number_format = "0"
        _wrap_cell(ws.cell(row=row, column=8, value=", ".join(p.targetKeywords or [])))
        _wrap_cell(ws.cell(row=row, column=9, value=", ".join(p.presentKeywords or [])))
        _wrap_cell(ws.cell(row=row, column=10, value=", ".join(p.missingKeywords or [])))
        findings_summary = "\n".join(
            f"[{f.severity.upper()}] {f.title}" for f in (p.findings or [])
        )
        _wrap_cell(ws.cell(row=row, column=11, value=findings_summary))

        reco = p.recommendation
        _wrap_cell(ws.cell(row=row, column=12, value=(reco.url if reco else "")))
        _wrap_cell(ws.cell(row=row, column=13, value=(reco.title if reco else "")))
        row += 1

    if row > 2:
        ws.auto_filter.ref = f"A1:M{row - 1}"
    ws.freeze_panes = "A2"


def _write_missing(ws: Worksheet, pages: list[MissingPage]) -> None:
    headers = ["URL recommandée", "Raison", "Volume / mois", "Priorité"]
    _write_header(ws, headers)
    for letter, width in zip(("A", "B", "C", "D"), (55, 80, 16, 14)):
        ws.column_dimensions[letter].width = width

    row = 2
    for p in pages:
        ws.cell(row=row, column=1, value=p.url).font = _MONO_FONT
        _wrap_cell(ws.cell(row=row, column=2, value=p.reason or ""))
        vol = p.estimatedSearchVolume or 0
        vc = ws.cell(row=row, column=3, value=vol if vol > 0 else "—")
        vc.font = _MONO_FONT
        vc.alignment = Alignment(horizontal="right")
        if isinstance(vc.value, int):
            vc.number_format = "#,##0"
        prio_cell = ws.cell(row=row, column=4, value=(p.priority or "medium").upper())
        prio_cell.fill = _PRIORITY_FILL.get(p.priority, PatternFill())
        prio_cell.font = Font(name="Inter", size=10, bold=True, color="18181B")
        prio_cell.alignment = Alignment(horizontal="center")
        row += 1

    if row > 2:
        ws.auto_filter.ref = f"A1:D{row - 1}"
    ws.freeze_panes = "A2"


def _write_technical_crawl(ws: Worksheet, tc) -> None:
    """Screaming-Frog-style crawl table — one row per visited URL."""
    headers = [
        "URL", "Code HTTP", "Indexable", "Raison non-indexable", "Profondeur",
        "Poids HTML (o)", "Mots", "Ratio texte/HTML", "Title (car.)",
        "Meta (car.)", "H1", "H2", "Liens internes", "Liens externes",
        "Images", "Images sans alt", "Type de page", "Problèmes",
    ]
    _write_header(ws, headers)
    widths = (55, 11, 11, 30, 11, 14, 9, 16, 12, 12, 7, 7, 14, 14, 9, 14, 16, 70)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2
    for r in tc.rows:
        ws.cell(row=row, column=1, value=r.url).font = _MONO_FONT
        sc = ws.cell(row=row, column=2, value=r.statusCode if r.statusCode is not None else "ERR")
        sc.font = _MONO_FONT
        sc.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value="Oui" if r.isIndexable else "Non").alignment = Alignment(horizontal="center")
        _wrap_cell(ws.cell(row=row, column=4, value=r.indexabilityReason or ""))
        dc = ws.cell(row=row, column=5, value=r.depth if r.depth is not None else "—")
        dc.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=r.htmlBytes).number_format = "#,##0"
        ws.cell(row=row, column=7, value=r.wordCount).number_format = "0"
        rc = ws.cell(row=row, column=8, value=round(r.textRatio, 3) if r.htmlBytes else "—")
        if r.htmlBytes:
            rc.number_format = "0.0%"
        ws.cell(row=row, column=9, value=r.titleLength).number_format = "0"
        ws.cell(row=row, column=10, value=r.metaDescLength).number_format = "0"
        ws.cell(row=row, column=11, value=r.h1Count).number_format = "0"
        ws.cell(row=row, column=12, value=r.h2Count).number_format = "0"
        ws.cell(row=row, column=13, value=r.internalLinksOut).number_format = "0"
        ws.cell(row=row, column=14, value=r.externalLinksOut).number_format = "0"
        ws.cell(row=row, column=15, value=r.imagesCount).number_format = "0"
        ws.cell(row=row, column=16, value=r.imagesWithoutAlt).number_format = "0"
        ws.cell(row=row, column=17, value=getattr(r, "pageType", "") or "")
        _wrap_cell(ws.cell(row=row, column=18, value="; ".join(r.issues) if r.issues else "OK"))
        row += 1

    if row > 2:
        ws.auto_filter.ref = f"A1:R{row - 1}"
    ws.freeze_panes = "A2"


def _write_a11y(ws: Worksheet, a) -> None:
    """Accessibility per-page scores + issues."""
    headers = ["URL", "Score a11y", "Problèmes détectés"]
    _write_header(ws, headers)
    for letter, width in zip(("A", "B", "C"), (55, 12, 100)):
        ws.column_dimensions[letter].width = width
    row = 2
    for p in a.pageScores:
        ws.cell(row=row, column=1, value=p.url).font = _MONO_FONT
        sc = ws.cell(row=row, column=2, value=p.score)
        sc.number_format = "0"
        sc.alignment = Alignment(horizontal="center")
        _wrap_cell(ws.cell(row=row, column=3, value="; ".join(p.issues) if p.issues else "—"))
        row += 1
    if row > 2:
        ws.auto_filter.ref = f"A1:C{row - 1}"
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Helpers


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    for col, label in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _THIN_BORDER
    ws.row_dimensions[1].height = 26


def _wrap_cell(cell) -> None:
    cell.font = _BODY_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = _THIN_BORDER


def _score_fill(score: int) -> PatternFill:
    if score < 40:
        return _SCORE_FILLS["critical"]
    if score < 60:
        return _SCORE_FILLS["warning"]
    if score < 80:
        return _SCORE_FILLS["info"]
    return _SCORE_FILLS["ok"]


def _score_fill_for_status(status: str) -> PatternFill:
    return {
        "critical": _SCORE_FILLS["critical"],
        "warning": _SCORE_FILLS["warning"],
        "improve": _SCORE_FILLS["warning"],
        "ok": _SCORE_FILLS["ok"],
    }.get(status, PatternFill())


def _progress_bar(score: int, width: int = 12) -> str:
    """ASCII progress bar that renders well in monospace on Excel."""
    filled = max(0, min(width, int(round(width * score / 100))))
    return "█" * filled + "░" * (width - filled)
